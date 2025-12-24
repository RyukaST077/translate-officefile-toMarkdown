"""Excel image detection utilities."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class ExcelImageDetectionResult:
    """Result of detecting images in an Excel workbook."""

    sheet_names: list[str]
    image_sheet_names: list[str]
    warnings: list[str]


def detect_excel_images(path: Path) -> ExcelImageDetectionResult:
    """Detect image-containing sheets for .xlsx files.

    Returns warnings when detection is skipped or fails.
    """
    extension = path.suffix.lower()
    if extension != ".xlsx":
        return ExcelImageDetectionResult(
            sheet_names=[],
            image_sheet_names=[],
            warnings=[".xlsx以外は画像検出をスキップしました。"],
        )

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive fallback
        return ExcelImageDetectionResult(
            sheet_names=[],
            image_sheet_names=[],
            warnings=[f"画像検出に失敗しました: {exc}"],
        )

    try:
        sheet_names = list(workbook.sheetnames)
        image_sheet_names: list[str] = []

        for worksheet in workbook.worksheets:
            images = _collect_sheet_images(worksheet)
            if images:
                image_sheet_names.append(worksheet.title)

        return ExcelImageDetectionResult(
            sheet_names=sheet_names,
            image_sheet_names=image_sheet_names,
            warnings=[],
        )
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()


def _collect_sheet_images(worksheet: object) -> Iterable[object]:
    images = getattr(worksheet, "_images", None)
    if images is None:
        images = getattr(worksheet, "images", None)
    if images is None:
        return []
    return images
