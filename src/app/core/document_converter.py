"""Document conversion utilities powered by MarkItDown."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from markitdown import MarkItDown
from openpyxl import load_workbook

from .excel_image_detector import ExcelImageDetectionResult, detect_excel_images
from .markdown_postprocessor import (
    MarkdownPostprocessResult,
    normalize_excel_markdown,
    remove_image_markdown,
)


@dataclass(frozen=True)
class ConversionResult:
    """Conversion result including markdown and warnings."""

    markdown: str
    warnings: list[str]


def convert_document(input_path: Path) -> ConversionResult:
    """Convert a document to Markdown and apply required post-processing."""
    converter = MarkItDown(enable_plugins=False)
    result = converter.convert(str(input_path))
    markdown = _extract_markdown(result)
    warnings: list[str] = []

    markdown = remove_image_markdown(markdown)

    extension = input_path.suffix.lower()
    if extension in {".xlsx", ".xls"}:
        image_result = detect_excel_images(input_path)
        warnings.extend(image_result.warnings)

        post_result = _postprocess_excel_markdown(
            markdown,
            image_result,
        )
        markdown = post_result.markdown
        warnings.extend(post_result.warnings)

        if extension == ".xlsx":
            warnings.extend(_detect_missing_formula_values(input_path))

    return ConversionResult(markdown=markdown, warnings=warnings)


def _extract_markdown(result: object) -> str:
    text = getattr(result, "text_content", None)
    if text is None:
        return str(result)
    return str(text)


def _postprocess_excel_markdown(
    markdown: str,
    image_result: ExcelImageDetectionResult,
) -> MarkdownPostprocessResult:
    return normalize_excel_markdown(
        markdown,
        getattr(image_result, "sheet_names", []),
        getattr(image_result, "image_sheet_names", []),
    )


def _detect_missing_formula_values(path: Path) -> list[str]:
    warnings: list[str] = []

    try:
        formula_book = load_workbook(path, data_only=False, read_only=True)
        value_book = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - defensive fallback
        return [f"数式結果の検出に失敗しました: {exc}"]

    missing_count = 0
    samples: list[str] = []
    truncated = False
    max_samples = 5
    max_missing = 50

    try:
        for sheet_name in formula_book.sheetnames:
            ws_formula = formula_book[sheet_name]
            ws_values = value_book[sheet_name]
            for row in ws_formula.iter_rows():
                for cell in row:
                    if cell.data_type != "f":
                        continue
                    value = ws_values[cell.coordinate].value
                    if value is None:
                        missing_count += 1
                        if len(samples) < max_samples:
                            samples.append(f"{sheet_name}!{cell.coordinate}")
                        if missing_count >= max_missing:
                            truncated = True
                            break
                if truncated:
                    break
            if truncated:
                break
    finally:
        for book in (formula_book, value_book):
            close = getattr(book, "close", None)
            if callable(close):
                close()

    if missing_count > 0:
        sample_text = ", ".join(samples)
        if truncated:
            warnings.append(
                "数式結果が取得できないセルが多数あります。"
                f"（例: {sample_text} ほか）"
            )
        else:
            warnings.append(
                "数式結果が取得できないセルがあります。"
                f"（例: {sample_text}）"
            )

    return warnings
